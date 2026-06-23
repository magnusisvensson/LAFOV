
import streamlit as st
import pandas as pd
from collections import defaultdict, deque
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO

st.set_page_config(layout="wide")

st.title("VFU-placeringssystem")

system_file = st.file_uploader("Översiktsfil", type=["xlsx"])
form_file = st.file_uploader("Formulärsvar", type=["xlsx"])

kull = st.number_input("Kull", value=26)
program = st.selectbox("Program", ["LAFOV","LAGRV","LGFRI"])


# =========================
# SESSION
# =========================
if "step" not in st.session_state:
    st.session_state.step = 1

if "student_regions" not in st.session_state:
    st.session_state.student_regions = {}

if "manual_assignments" not in st.session_state:
    st.session_state.manual_assignments = {}

if "excel_data" not in st.session_state:
    st.session_state.excel_data = None


# =========================
# REGION
# =========================
def get_region(text):
    t = str(text).lower()
    if "oskarshamn" in t:
        return "Oskarshamn"
    if any(x in t for x in ["karlskrona","ronneby","rödeby"]):
        return "Karlskrona"
    return "Kalmar"


if system_file and form_file:

    # =========================
    # SKOLOR
    # =========================
    skolor = pd.read_excel(system_file)
    skolor.columns = skolor.columns.str.strip()
    skolor["Skolenhet"] = skolor["Skolenhet"].astype(str).str.strip()

    skolor = skolor[
        (
            (skolor["Kull"] == kull) |
            (skolor["Kull"].astype(str).str.upper() == "VAKANT")
        ) &
        (skolor["Inriktning"].str.upper() == program)
    ].copy()

    skolor["Region"] = skolor["Partnerområde"].apply(get_region)

    kap = {}
    for _, r in skolor.iterrows():
        try:
            kap[r["Skolenhet"]] = int(float(r["Antal platser"]))
        except:
            kap[r["Skolenhet"]] = 2


    # =========================
    # STUDENTER
    # =========================
    students = pd.read_excel(form_file, sheet_name="Data")
    students.columns = students.columns.str.strip()

    fn=[c for c in students.columns if "förnamn" in c.lower()][0]
    ln=[c for c in students.columns if "efternamn" in c.lower()][0]
    bost=[c for c in students.columns if "bostads" in c.lower()][0]

    students["Student"]=(students[fn]+" "+students[ln]).str.strip()
    students["Ort"]=students[bost]


    # =========================
    # STEG 1
    # =========================
    if st.session_state.step == 1:

        st.header("1. Region")

        temp={}

        for _,row in students.iterrows():

            name=row["Student"]
            ort=str(row["Ort"]).lower()

            if "kalmar" in ort:
                region="Kalmar"
            elif "karlskrona" in ort:
                region="Karlskrona"
            elif "oskarshamn" in ort:
                region="Oskarshamn"
            else:
                region=st.selectbox(
                    f"{name} ({row['Ort']})",
                    ["Kalmar","Karlskrona","Oskarshamn"],
                    key=name
                )

            temp[name]=region

        if st.button("✅ Bekräfta"):
            st.session_state.student_regions=temp
            st.session_state.step=2
            st.rerun()


    # =========================
    # STEG 2
    # =========================
    if st.session_state.step == 2:

        usage = defaultdict(lambda:{
            "År1":0,"År2":0,"År3":0,"År4":0
        })

        # ✅ KAPACITETSSTYRD FUNKTION
        def best_school(skolor_list, year):

            def load(sk):
                return usage[sk][year] / kap.get(sk,1)

            return min(skolor_list, key=load)

        results=[]

        for _,s in students.iterrows():

            student=s["Student"]
            region=st.session_state.student_regions.get(student)

            skolor_r = skolor[skolor["Region"]==region]["Skolenhet"].tolist()

            if not skolor_r:
                continue

            # ✅ KAPACITETSVAL
            A = best_school(skolor_r, "År1")

            rest1 = [sk for sk in skolor_r if sk != A]
            B = best_school(rest1, "År2") if rest1 else A

            rest2 = [sk for sk in rest1 if sk != B]
            C = best_school(rest2, "År4") if rest2 else B

            if program=="LGFRI":
                y1,y2,y3,y4=A,A,B,""
            else:
                if region=="Kalmar":
                    y1,y2,y3,y4=A,B,B,C
                else:
                    y1,y2,y3,y4=A,B,A,B

            # ✅ UPDATE usage (VIKTIGT)
            usage[A]["År1"] += 1
            usage[B]["År2"] += 1
            usage[B]["År3"] += 1
            if y4:
                usage[C]["År4"] += 1

            results.append({
                "Student":student,
                "Ort":s["Ort"],
                "Region":region,
                "År1":y1,
                "År2":y2,
                "År3":y3,
                "År4":y4
            })

        df=pd.DataFrame(results)


        # =========================
        # MANUELL EDITOR (OFÖRÄNDRAD)
        # =========================
        st.header("📝 Justera placering manuellt")

        student_name = st.selectbox("Välj student", df["Student"])

        row = df[df["Student"]==student_name].iloc[0]
        region=row["Region"]

        skolor_r = skolor[skolor["Region"]==region]["Skolenhet"].tolist()

        if student_name not in st.session_state.manual_assignments:
            st.session_state.manual_assignments[student_name] = {
                "År1":row["År1"],
                "År2":row["År2"],
                "År3":row["År3"],
                "År4":row["År4"],
            }

        current=st.session_state.manual_assignments[student_name]

        col1,col2,col3,col4=st.columns(4)

        current["År1"]=col1.selectbox("År1",skolor_r,index=skolor_r.index(current["År1"]))
        current["År2"]=col2.selectbox("År2",skolor_r,index=skolor_r.index(current["År2"]))
        current["År3"]=col3.selectbox("År3",skolor_r,index=skolor_r.index(current["År3"]))

        if current["År4"]:
            current["År4"]=col4.selectbox("År4",skolor_r,index=skolor_r.index(current["År4"]))
        else:
            current["År4"]=col4.selectbox("År4",[""]+skolor_r)

        if st.button("✅ Spara ändring"):
            st.session_state.manual_assignments[student_name]=current
            st.success("Sparat")


        # =========================
        # SLÅ IHOP (OFÖRÄNDRAT)
        # =========================
        export_rows=[]

        for _,r in df.iterrows():

            student=r["Student"]

            if student in st.session_state.manual_assignments:
                merged={
                    "Student":student,
                    "Ort":r["Ort"],
                    "Region":r["Region"],
                    "År1":st.session_state.manual_assignments[student]["År1"],
                    "År2":st.session_state.manual_assignments[student]["År2"],
                    "År3":st.session_state.manual_assignments[student]["År3"],
                    "År4":st.session_state.manual_assignments[student]["År4"],
                }
            else:
                merged=r

            export_rows.append(merged)

        export_df=pd.DataFrame(export_rows)


        # =========================
        # EXCEL (OFÖRÄNDRAD)
        # =========================
        def build_excel(df):

    output = BytesIO()
    wb = Workbook()

    thin = Side(style="thin")
    medium = Side(style="medium")

    box = Border(top=medium, left=medium, right=medium, bottom=medium)
    inner = Border(left=thin, right=thin)

    fill_region = PatternFill("solid","BDD7EE")
    fill_header = PatternFill("solid","E7F3FF")

    # -------- BLAD 1 --------
    ws = wb.active
    ws.title = "Placeringar"

    row_i = 1

    for region in ["Kalmar","Oskarshamn","Karlskrona"]:

        # ----- REGION -----
        ws.cell(row_i,1,region)
        ws.merge_cells(start_row=row_i,start_column=1,end_row=row_i,end_column=5)

        for c in range(1,6):
            ws.cell(row_i,c).fill = fill_region

        ws.cell(row_i,1).alignment = Alignment(horizontal="center")

        row_i += 2  # luft

        # ----- SKOLOR -----
        for skola in skolor[skolor["Region"]==region]["Skolenhet"]:

            start_block = row_i

            # skolrubrik
            ws.cell(row_i,1,f"{skola} (max {kap[skola]})")
            ws.merge_cells(start_row=row_i,start_column=1,end_row=row_i,end_column=5)

            ws.cell(row_i,1).fill = fill_header
            row_i += 1

            # kolumnrubrik
            headers = ["","År1","År2","År3","År4"]
            ws.append(headers)

            for c in range(2,6):
                ws.cell(row_i,c).alignment = Alignment(horizontal="center")

            row_i += 1

            # data
            subset = df[
                (df["År1"]==skola) |
                (df["År2"]==skola) |
                (df["År3"]==skola) |
                (df["År4"]==skola)
            ]

            for _, r in subset.iterrows():
                ws.append([
                    "",
                    r["Student"] if r["År1"]==skola else "",
                    r["Student"] if r["År2"]==skola else "",
                    r["Student"] if r["År3"]==skola else "",
                    r["Student"] if r["År4"]==skola else "",
                ])
                row_i += 1

            end_block = row_i - 1

            # ✅ RAM RUNT HELA TABELLEN
            for r_i in range(start_block, end_block+1):
                for c_i in range(1,6):

                    cell = ws.cell(r_i,c_i)

                    # ytterram
                    border = Border(
                        left = medium if c_i==1 else thin,
                        right = medium if c_i==5 else thin,
                        top = medium if r_i==start_block else thin,
                        bottom = medium if r_i==end_block else thin
                    )
                    cell.border = border

            row_i += 2  # luft mellan skolor

    # -------- BLAD 2 --------
    ws2 = wb.create_sheet("Studenter")

    ws2.append(["Student","Ort","Region","År1","År2","År3","År4"])

    for _, r in df.iterrows():
        ws2.append(list(r))

    # -------- BLAD 3 --------
    ws3 = wb.create_sheet("Kontroll")

    ws3.append(["Student","Antal skolor","Status"])

    fill_ok = PatternFill("solid","C6EFCE")
    fill_warn = PatternFill("solid","FFEB9C")

    for _, r in df.iterrows():

        skolset = {r["År1"],r["År2"],r["År3"],r["År4"]}
        skolset.discard("")
        antal = len(skolset)

        if antal < 2:
            status = "⚠"
            color = fill_warn
        else:
            status = "OK"
            color = fill_ok

        ws3.append([r["Student"], antal, status])

        ws3.cell(ws3.max_row,3).fill = color

    # -------- AUTOBREDD --------
    for sheet in wb.worksheets:
        for col in sheet.columns:
            length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
            sheet.column_dimensions[get_column_letter(col[0].column)].width = length + 3

    wb.save(output)
    output.seek(0)

    return output



        st.session_state.excel_data = build_excel(export_df)

        if st.session_state.excel_data:
            st.download_button(
                "⬇️ Ladda ner Excel",
                data=st.session_state.excel_data,
                file_name="kull_resultat.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
